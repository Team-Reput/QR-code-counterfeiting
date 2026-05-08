"""
qr_generator.py
===============
Cryptographic QR Code Generator for the Reput Anti-Counterfeiting System.

Usage
-----
  python qr_generator.py --input products.xlsx [OPTIONS]
  python qr_generator.py --input products.csv  [OPTIONS]

Required argument
  --input PATH        Excel (.xlsx) or CSV (.csv) file with product data.
                      Every column is included in the signed payload.
                      The file MUST contain a 'batch_id' column (used as
                      the lookup key in otp_codes.csv by the backend).

Optional arguments
  --privkey PATH      Ed25519 private key PEM file.
                      Default: private_key.pem (next to this script).
                      If the file does not exist a new keypair is generated
                      and BOTH private_key.pem and public_key.pem are written.
  --pubkey  PATH      Ed25519 public key PEM output path.
                      Default: public_key.pem (next to this script).
  --url     URL       Base verification URL embedded in each QR code.
                      The token is appended as ?token=<signature>.
                      Default: http://localhost:5000/verify
  --output  DIR       Folder to write QR PNG images.
                      Default: ./qr_output/
  --otp-csv PATH      CSV file to append OTP rows into.
                      Default: otp_codes.csv (next to this script).
  --sheet   NAME      Sheet name when reading an Excel file.
                      Default: first sheet.
  --append            If set, new rows are appended to an existing otp-csv.
                      If not set, the otp-csv is REPLACED on each run.

The generated otp_codes.csv columns are:
  batch_id, signature_token, otp_code, token, owner_name, owner_phone,
  device_fingerprint, device_make, device_model, claimed_at

These are 100% compatible with backend_verification.py without any changes.
"""

import argparse
import base64
import csv
import json
import os
import random
import string
import struct
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import qrcode
from PIL import Image, ImageDraw

from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey


# ================================================================
#                       CONSTANTS
#   Must stay in sync with backend_verification.py exactly.
# ================================================================
QR_BOX_SIZE = int(os.getenv("QR_BOX_SIZE", "10"))
QR_BORDER   = int(os.getenv("QR_BORDER",   "2"))
QR_ERROR_CORRECTION = qrcode.constants.ERROR_CORRECT_H

TOP_STRIP_THICKNESS   = 4
RIGHT_STRIP_THICKNESS = 4


# ================================================================
#                    JSON CANONICAL ENCODER
#   Must match backend_verification.py: sort_keys, separators
# ================================================================
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj: Any) -> Any:
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)


def canonical_product_json(row: Dict[str, Any]) -> bytes:
    """
    Produce the canonical UTF-8 JSON bytes for a product row.
    Must be identical to the function in backend_verification.py.
    """
    return json.dumps(
        row,
        cls=DateTimeEncoder,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


# ================================================================
#                    KEY MANAGEMENT
# ================================================================
def load_or_generate_private_key(
    privkey_path: str, pubkey_path: str
) -> Ed25519PrivateKey:
    """
    Load an existing Ed25519 private key from PEM, or generate a fresh
    keypair, saving both private_key.pem and public_key.pem.
    """
    if os.path.exists(privkey_path):
        print(f"[INFO] Loading private key from: {privkey_path}")
        with open(privkey_path, "rb") as f:
            private_key = serialization.load_pem_private_key(f.read(), password=None)
        return private_key

    # Generate new keypair
    print(f"[INFO] No private key found at '{privkey_path}'. Generating new Ed25519 keypair…")
    private_key = Ed25519PrivateKey.generate()

    # Save private key
    priv_pem = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption(),
    )
    with open(privkey_path, "wb") as f:
        f.write(priv_pem)
    print(f"[INFO] Private key saved -> {privkey_path}")

    # Save matching public key (so backend_verification.py can verify)
    pub_pem = private_key.public_key().public_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PublicFormat.SubjectPublicKeyInfo,
    )
    with open(pubkey_path, "wb") as f:
        f.write(pub_pem)
    print(f"[INFO] Public key saved  -> {pubkey_path}  (update backend PUBLIC_KEY_PATH if needed)")

    return private_key


# ================================================================
#                    SIGNING
# ================================================================
def sign_ed25519_b64url(private_key: Ed25519PrivateKey, data: bytes) -> str:
    """
    Sign `data` with the private key and return a URL-safe base64 string
    without padding, exactly matching what the backend expects.
    """
    signature: bytes = private_key.sign(data)
    b64 = base64.urlsafe_b64encode(signature).decode("utf-8")
    return b64.rstrip("=")   # strip padding (backend re-pads on verify)


# ================================================================
#                    OTP GENERATION
# ================================================================
_OTP_CHARS = string.ascii_uppercase + string.digits


def generate_otp(length: int = 6) -> str:
    """Generate a random OTP that matches the format in otp_codes.csv."""
    return "".join(random.choices(_OTP_CHARS, k=length))


# ================================================================
#                    QR IMAGE HELPERS
#   Copied verbatim from backend_verification.py so output is
#   pixel-identical and the micro-pattern verify passes.
# ================================================================
def make_qr_image(payload_text: str) -> Image.Image:
    qr = qrcode.QRCode(
        version=None,
        error_correction=QR_ERROR_CORRECTION,
        box_size=QR_BOX_SIZE,
        border=QR_BORDER,
    )
    qr.add_data(payload_text)
    qr.make(fit=True)
    return qr.make_image(fill_color="black", back_color="white").convert("RGB")


def apply_micro_pattern_top_and_right(qr_img: Image.Image) -> Image.Image:
    """
    Embed the micro-pattern (white dots in top + right strips on black modules).
    Logic is identical to backend_verification.py so the verifier passes.
    """
    img  = qr_img.convert("RGB")
    draw = ImageDraw.Draw(img)

    w, _ = img.size
    modules = w // QR_BOX_SIZE

    def in_quiet_zone(mx: int, my: int) -> bool:
        return (
            mx < QR_BORDER or
            my < QR_BORDER or
            mx >= (modules - QR_BORDER) or
            my >= (modules - QR_BORDER)
        )

    def in_finder_zone(mx: int, my: int) -> bool:
        if mx < (QR_BORDER + 9) and my < (QR_BORDER + 9):
            return True
        if mx >= (modules - QR_BORDER - 9) and my < (QR_BORDER + 9):
            return True
        if mx < (QR_BORDER + 9) and my >= (modules - QR_BORDER - 9):
            return True
        return False

    timing_row = QR_BORDER + 6
    timing_col = QR_BORDER + 6

    dot_r  = max(2, QR_BOX_SIZE // 4)
    offset = max(1, QR_BOX_SIZE // 3)

    top_y_start   = QR_BORDER
    top_y_end     = min(modules - QR_BORDER, QR_BORDER + TOP_STRIP_THICKNESS)
    right_x_start = max(QR_BORDER, modules - QR_BORDER - RIGHT_STRIP_THICKNESS)
    right_x_end   = modules - QR_BORDER

    for my in range(QR_BORDER, modules - QR_BORDER):
        for mx in range(QR_BORDER, modules - QR_BORDER):
            if in_quiet_zone(mx, my) or in_finder_zone(mx, my):
                continue
            if mx == timing_col or my == timing_row:
                continue

            in_top_strip   = top_y_start <= my < top_y_end
            in_right_strip = right_x_start <= mx < right_x_end
            if not (in_top_strip or in_right_strip):
                continue

            x0 = mx * QR_BOX_SIZE
            y0 = my * QR_BOX_SIZE
            cx = x0 + QR_BOX_SIZE // 2
            cy = y0 + QR_BOX_SIZE // 2

            r, g, b = img.getpixel((cx, cy))
            if (r, g, b) == (0, 0, 0):            # only black modules get a dot
                dx = cx - offset // 2
                dy = cy + offset // 2
                draw.ellipse(
                    (dx - dot_r, dy - dot_r, dx + dot_r, dy + dot_r),
                    fill=(255, 255, 255),
                )

    return img


# ================================================================
#              INVISIBLE LSB WATERMARK
#
#  Technique: Least Significant Bit (LSB) steganography.
#  The watermark JSON is encoded as bits and written into the LSB
#  of the Red channel of every pixel, row by row.
#
#  Why it does NOT affect backend verification:
#   - CSV token lookup  = pure string comparison, no pixels read.
#   - Micro-pattern check uses threshold  r > 200.
#     White pixels: 255 -> 254 (still > 200).  Black: 0 -> 1 (< 200).
#  Why it is invisible:
#   - A +/-1 change in an 8-bit channel value is imperceptible.
#  Format: first 4 bytes = payload length (big-endian uint32),
#          followed by UTF-8 JSON payload bytes.
# ================================================================
WATERMARK_BRAND_DEFAULT = "Reput"


def embed_watermark(img: Image.Image, payload: Dict[str, Any]) -> Image.Image:
    """
    Embed a JSON payload invisibly into the image using LSB steganography.
    Returns a new image; the original is not mutated.

    The payload is serialised to compact JSON UTF-8, length-prefixed with
    a 4-byte big-endian unsigned int, then each bit is written into the
    LSB of the R channel of successive pixels (left-to-right, top-to-bottom).

    Maximum embeddable payload = (width * height // 8) - 4 bytes.
    For a typical 370x370 QR image that is ~17,000 bytes — far more than enough.
    """
    img    = img.copy().convert("RGB")
    pixels = img.load()
    w, h   = img.size

    # Serialise payload to compact UTF-8 JSON
    payload_bytes: bytes = json.dumps(
        payload, separators=(",", ":"), sort_keys=True
    ).encode("utf-8")
    length: int = len(payload_bytes)

    # 4-byte big-endian length prefix + data
    raw: bytes = struct.pack(">I", length) + payload_bytes

    # Flatten to bit list (MSB first per byte)
    bits: List[int] = []
    for byte in raw:
        for bit_pos in range(7, -1, -1):
            bits.append((byte >> bit_pos) & 1)

    max_bits = w * h
    if len(bits) > max_bits:
        raise ValueError(
            f"Watermark payload ({length} bytes) is too large for this "
            f"{w}x{h} image (max {max_bits // 8 - 4} bytes)."
        )

    # Write each bit into the LSB of the R channel pixel by pixel
    bit_idx = 0
    for y in range(h):
        for x in range(w):
            if bit_idx >= len(bits):
                break
            r, g, b = pixels[x, y]
            r = (r & 0xFE) | bits[bit_idx]   # clear LSB then set new value
            pixels[x, y] = (r, g, b)
            bit_idx += 1
        if bit_idx >= len(bits):
            break

    return img


def extract_watermark(img: Image.Image) -> Optional[Dict[str, Any]]:
    """
    Extract and return the watermark payload dict from an LSB-watermarked image.
    Returns None if no valid watermark is found.

    Standalone usage example::

        from PIL import Image
        from qr_generator import extract_watermark
        data = extract_watermark(Image.open("batch_1001.png"))
        # {'batch_id': '1001', 'generated_at': '2026-05-03T14:08:05', 'brand': 'Reput'}
    """
    img    = img.convert("RGB")
    pixels = img.load()
    w, h   = img.size
    total  = w * h

    def _read_n_bits(start_pixel: int, n: int) -> List[int]:
        result: List[int] = []
        idx = start_pixel
        while len(result) < n and idx < total:
            x, y = idx % w, idx // w
            r, _, _ = pixels[x, y]
            result.append(r & 1)
            idx += 1
        return result

    def _bits_to_int(bits: List[int]) -> int:
        val = 0
        for b in bits:
            val = (val << 1) | b
        return val

    def _bits_to_bytes(bits: List[int]) -> bytes:
        out = bytearray()
        for i in range(0, len(bits) - 7, 8):
            out.append(_bits_to_int(bits[i: i + 8]))
        return bytes(out)

    # --- Read 32-bit length header (pixels 0-31) ---
    header_bits = _read_n_bits(0, 32)
    if len(header_bits) < 32:
        return None
    length = _bits_to_int(header_bits)

    # Sanity: reject impossible lengths
    if length <= 0 or length > (total // 8) - 4:
        return None

    # --- Read payload bits (pixels 32 onward) ---
    payload_bits  = _read_n_bits(32, length * 8)
    payload_bytes = _bits_to_bytes(payload_bits)

    try:
        return json.loads(payload_bytes.decode("utf-8"))
    except Exception:
        return None


# ================================================================
#                    PRODUCT FILE PARSER
# ================================================================
def _coerce_row(row: Dict[str, str]) -> Dict[str, Any]:
    """
    Try to coerce numeric strings to int/float so that the canonical
    JSON matches the type profile the backend would produce from the DB.
    Strings that are not numeric are left as-is.
    """
    coerced: Dict[str, Any] = {}
    for k, v in row.items():
        if v is None or (isinstance(v, str) and v.strip() == ""):
            coerced[k] = None
            continue
        s = str(v).strip()
        # Try int first
        try:
            coerced[k] = int(s)
            continue
        except ValueError:
            pass
        # Try float
        try:
            coerced[k] = float(s)
            continue
        except ValueError:
            pass
        # Leave as string
        coerced[k] = s
    return coerced


def parse_product_file(
    path: str, sheet_name: Optional[str] = None
) -> List[Dict[str, Any]]:
    """
    Parse an Excel (.xlsx) or CSV (.csv) product file.
    Returns a list of row dicts with all columns included.
    Empty rows are skipped.
    """
    ext = Path(path).suffix.lower()

    if ext in (".xlsx", ".xls", ".xlsm"):
        return _parse_excel(path, sheet_name)
    elif ext == ".csv":
        return _parse_csv(path)
    else:
        raise ValueError(
            f"Unsupported file type: '{ext}'. "
            "Please provide a .xlsx, .xls, or .csv file."
        )


def _parse_excel(path: str, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl is required for Excel files.  Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(next(rows_iter))]
    except StopIteration:
        return []

    products = []
    for raw_row in rows_iter:
        row_dict = {headers[i]: raw_row[i] for i in range(len(headers))}
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row_dict.values()):
            continue
        # Convert types
        coerced = {}
        for k, v in row_dict.items():
            if v is None:
                coerced[k] = None
            elif isinstance(v, (int, float, bool, datetime, date)):
                coerced[k] = v       # keep native type
            else:
                s = str(v).strip()
                try:
                    coerced[k] = int(s)
                except ValueError:
                    try:
                        coerced[k] = float(s)
                    except ValueError:
                        coerced[k] = s if s != "" else None
        products.append(coerced)

    return products


def _parse_csv(path: str) -> List[Dict[str, Any]]:
    products = []
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for raw_row in reader:
            # Strip whitespace from keys and values
            row = {k.strip(): v.strip() if isinstance(v, str) else v
                   for k, v in raw_row.items()}
            # Skip completely empty rows
            if all(v == "" or v is None for v in row.values()):
                continue
            products.append(_coerce_row(row))
    return products


# ================================================================
#                    OTP CSV HELPERS
# ================================================================
OTP_CSV_FIELDNAMES = [
    "batch_id",
    "signature_token",
    "otp_code",
    "token",
    "owner_name",
    "owner_phone",
    "device_fingerprint",
    "device_make",
    "device_model",
    "claimed_at",
]


def _read_existing_csv(otp_csv_path: str) -> List[Dict[str, str]]:
    """Read existing rows from otp_codes.csv (for append mode)."""
    if not os.path.exists(otp_csv_path):
        return []
    with open(otp_csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)


def write_otp_csv(
    otp_csv_path: str,
    new_rows: List[Dict[str, str]],
    append: bool = False,
) -> None:
    """
    Write (or append) new OTP rows to the CSV.
    Columns match backend_verification.py expectations exactly.
    """
    existing: List[Dict[str, str]] = []
    if append:
        existing = _read_existing_csv(otp_csv_path)

    all_rows = existing + new_rows

    with open(otp_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OTP_CSV_FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_rows)


# ================================================================
#                    MAIN GENERATOR LOGIC
# ================================================================
def generate_qr_for_product(
    product: Dict[str, Any],
    private_key: Ed25519PrivateKey,
    base_url: str,
    output_dir: str,
    row_index: int,
    watermark: bool = True,
    brand: str = WATERMARK_BRAND_DEFAULT,
) -> Dict[str, str]:
    """
    Generate a signed, micro-patterned QR code for a single product row.
    Optionally embeds an invisible LSB watermark (batch_id + timestamp + brand).
    Returns the otp_codes.csv row dict for this product.
    """
    # 1. Build canonical JSON bytes (exactly as backend does it)
    data_bytes = canonical_product_json(product)

    # 2. Sign with Ed25519 private key → base64url token (no padding)
    signature_token = sign_ed25519_b64url(private_key, data_bytes)

    # 3. Build the QR payload URL  (same structure backend extracts token from)
    qr_url = f"{base_url.rstrip('/')}?token={signature_token}"

    # 4. Generate the base QR image
    qr_base = make_qr_image(qr_url)

    # 5. Apply micro-pattern (identical to backend verifier)
    qr_final = apply_micro_pattern_top_and_right(qr_base.copy())

    # 6. Embed invisible LSB watermark (does NOT affect QR scan or backend auth)
    batch_id = str(product.get("batch_id", "") or product.get("batchid", "") or row_index)
    if watermark:
        wm_payload = {
            "batch_id":     batch_id,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "brand":        brand,
        }
        try:
            qr_final = embed_watermark(qr_final, wm_payload)
        except Exception as wm_err:
            print(f"  [WARN] Watermark skipped for batch {batch_id}: {wm_err}")

    # 7. Determine output filename
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in str(batch_id))
    png_path = os.path.join(output_dir, f"batch_{safe_name}.png")

    # 8. Save QR PNG
    qr_final.save(png_path)

    # 9. Generate OTP
    otp = generate_otp()

    wm_label = "(+watermark)" if watermark else ""
    print(f"  [OK]  batch_id={batch_id:>6}  OTP={otp}  -> {png_path} {wm_label}")

    # 10. Return the CSV row (owner fields left blank — claimed later via app)
    return {
        "batch_id":          batch_id,
        "signature_token":   signature_token,
        "otp_code":          otp,
        "token":             "",          # filled when owner scans & claims
        "owner_name":        "",
        "owner_phone":       "",
        "device_fingerprint": "",
        "device_make":       "",
        "device_model":      "",
        "claimed_at":        "",
    }


def run(args: argparse.Namespace) -> None:
    # ── Resolve paths ──────────────────────────────────────────
    script_dir   = os.path.dirname(os.path.abspath(__file__))
    privkey_path = os.path.abspath(args.privkey)
    pubkey_path  = os.path.abspath(args.pubkey)
    output_dir   = os.path.abspath(args.output)
    otp_csv_path = os.path.abspath(args.otp_csv)
    input_path   = os.path.abspath(args.input)

    # ── Validate input file ────────────────────────────────────
    if not os.path.exists(input_path):
        print(f"[ERROR] Input file not found: {input_path}")
        sys.exit(1)

    # ── Load / generate keypair ────────────────────────────────
    private_key = load_or_generate_private_key(privkey_path, pubkey_path)

    # ── Parse product file ─────────────────────────────────────
    print(f"\n[INFO] Parsing product file: {input_path}")
    products = parse_product_file(input_path, sheet_name=args.sheet)

    if not products:
        print("[WARN] No product rows found in the file. Nothing generated.")
        sys.exit(0)

    print(f"[INFO] Found {len(products)} product(s). Columns: {list(products[0].keys())}")

    # Warn if batch_id column is missing (needed for backend OTP lookup)
    sample_keys = set(products[0].keys())
    if "batch_id" not in sample_keys and "batchid" not in sample_keys:
        print(
            "[WARN] No 'batch_id' column detected. The backend uses 'batch_id' to look up "
            "OTP codes. The generated CSV will use the row number as batch_id.\n"
            "       Consider adding a 'batch_id' column to your input file."
        )

    # ── Prepare output folder ──────────────────────────────────
    os.makedirs(output_dir, exist_ok=True)
    print(f"[INFO] Output folder: {output_dir}")

    # ── Generate QR codes ──────────────────────────────────────
    print(f"\n[INFO] Generating {len(products)} QR code(s)…\n")
    otp_rows: List[Dict[str, str]] = []
    for idx, product in enumerate(products, start=1):
        try:
            csv_row = generate_qr_for_product(
                product=product,
                private_key=private_key,
                base_url=args.url,
                output_dir=output_dir,
                row_index=idx,
                watermark=not args.no_watermark,
                brand=args.brand,
            )
            otp_rows.append(csv_row)
        except Exception as exc:
            print(f"  [ERR] Row {idx}: {exc}")

    # ── Write otp_codes.csv ────────────────────────────────────
    write_otp_csv(otp_csv_path, otp_rows, append=args.append)
    mode = "appended to" if args.append else "written to"
    print(f"\n[INFO] OTP CSV {mode}: {otp_csv_path}")
    print(f"[DONE] {len(otp_rows)} QR code(s) generated in: {output_dir}\n")


# ================================================================
#                    CLI ENTRY POINT
# ================================================================
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="qr_generator.py",
        description=(
            "Generate Ed25519-signed, micro-patterned QR codes from an Excel or CSV "
            "product list, compatible with backend_verification.py."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    parser.add_argument(
        "--input", "-i",
        required=True,
        metavar="PATH",
        help="Excel (.xlsx) or CSV (.csv) file with product data.",
    )
    parser.add_argument(
        "--url", "-u",
        default="https://qr.reputinfo.life/",
        metavar="URL",
        help=(
            "Base verification URL embedded in each QR code. "
            "Default: http://localhost:5000/verify"
        ),
    )
    parser.add_argument(
        "--privkey",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "private_key.pem"),
        metavar="PATH",
        help=(
            "Ed25519 private key PEM file. "
            "If not found, a new keypair is generated. "
            "Default: private_key.pem (same folder as this script)."
        ),
    )
    parser.add_argument(
        "--pubkey",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "public_key.pem"),
        metavar="PATH",
        help=(
            "Ed25519 public key PEM output path (written only when keypair is generated). "
            "Default: public_key.pem (same folder as this script)."
        ),
    )
    parser.add_argument(
        "--output", "-o",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "qr_output"),
        metavar="DIR",
        help="Folder to save QR PNG images. Default: ./qr_output/",
    )
    parser.add_argument(
        "--otp-csv",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "otp_codes.csv"),
        metavar="PATH",
        help="OTP CSV file to write / append. Default: otp_codes.csv",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        metavar="NAME",
        help="Sheet name when reading an Excel file. Default: first sheet.",
    )
    parser.add_argument(
        "--append",
        action="store_true",
        help=(
            "Append new rows to an existing otp-csv instead of replacing it. "
            "Useful for batching across multiple runs."
        ),
    )
    parser.add_argument(
        "--no-watermark",
        action="store_true",
        help="Disable the invisible LSB watermark embedded in each QR PNG.",
    )
    parser.add_argument(
        "--brand",
        default=WATERMARK_BRAND_DEFAULT,
        metavar="NAME",
        help=(
            f"Brand name embedded in the invisible watermark payload. "
            f"Default: '{WATERMARK_BRAND_DEFAULT}'."
        ),
    )

    return parser


if __name__ == "__main__":
    parser = build_parser()
    args   = parser.parse_args()
    run(args)
